"""
export_excel.py - Export company records from DB to Excel (.xlsx).
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook

from config import CSV_OUTPUT_PATH
from database import get_all_companies

FIELDNAMES = [
    "id",
    "source",
    "company_name",
    "country",
    "city",
    "website_url",
    "phone",
    "email",
    "scraped_at",
]


def _dedupe_rows(rows: list[dict]) -> list[dict]:
    deduped = []
    seen = set()
    for row in rows:
        key = (
            (row.get("source") or "").strip().lower(),
            (row.get("company_name") or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def _normalize_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value if value is not None else ""


def _write_sheet(ws, rows: list[dict]):
    ws.append(FIELDNAMES)
    for row in rows:
        ws.append([_normalize_value(row.get(k, "")) for k in FIELDNAMES])


def _safe_save(workbook: Workbook, path: str) -> str:
    try:
        workbook.save(path)
        return path
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        root, ext = os.path.splitext(path)
        fallback = f"{root}_{stamp}{ext}"
        print(f"[XLSX] File locked: {os.path.abspath(path)}")
        print(f"[XLSX] Writing to fallback file instead: {os.path.abspath(fallback)}")
        workbook.save(fallback)
        return fallback


def export_to_excel(output_path: str | None = None):
    """
    Export:
    - one combined workbook: output/companies.xlsx with sheets ALL, CLUTCH, GOODFIRMS
    - two source workbooks: output/companies_clutch.xlsx and output/companies_goodfirms.xlsx
    """
    rows = get_all_companies()
    if not rows:
        print("[XLSX] No data found in database.")
        return

    deduped = _dedupe_rows(rows)
    clutch_rows = [r for r in deduped if (r.get("source") or "").strip().lower() == "clutch"]
    goodfirms_rows = [r for r in deduped if (r.get("source") or "").strip().lower() == "goodfirms"]

    base = output_path or CSV_OUTPUT_PATH
    out_dir = os.path.dirname(base) or "."
    os.makedirs(out_dir, exist_ok=True)

    combined_path = os.path.join(out_dir, "companies.xlsx")
    clutch_path = os.path.join(out_dir, "companies_clutch.xlsx")
    goodfirms_path = os.path.join(out_dir, "companies_goodfirms.xlsx")

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "ALL"
    _write_sheet(ws_all, deduped)
    ws_clutch = wb.create_sheet("CLUTCH")
    _write_sheet(ws_clutch, clutch_rows)
    ws_goodfirms = wb.create_sheet("GOODFIRMS")
    _write_sheet(ws_goodfirms, goodfirms_rows)
    actual_combined = _safe_save(wb, combined_path)

    wb_clutch = Workbook()
    _write_sheet(wb_clutch.active, clutch_rows)
    wb_clutch.active.title = "CLUTCH"
    actual_clutch = _safe_save(wb_clutch, clutch_path)

    wb_goodfirms = Workbook()
    _write_sheet(wb_goodfirms.active, goodfirms_rows)
    wb_goodfirms.active.title = "GOODFIRMS"
    actual_goodfirms = _safe_save(wb_goodfirms, goodfirms_path)

    print(f"[XLSX] Combined: {len(deduped)} rows -> {os.path.abspath(actual_combined)}")
    print(f"[XLSX] Clutch: {len(clutch_rows)} rows -> {os.path.abspath(actual_clutch)}")
    print(f"[XLSX] GoodFirms: {len(goodfirms_rows)} rows -> {os.path.abspath(actual_goodfirms)}")


if __name__ == "__main__":
    export_to_excel()
